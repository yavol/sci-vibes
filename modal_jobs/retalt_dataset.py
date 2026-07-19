"""Prepare a leakage-resistant RETALT1 benchmark in a private Modal volume.

The research branches receive only ``public/train.csv`` and
``public/validation.csv``. The raw workbook, split manifest, and hidden labels
remain under ``private/`` and are never copied into branch repositories.
"""

from __future__ import annotations

import hashlib
import json
import re
from pathlib import Path
from urllib.request import urlopen

import modal


APP_NAME = "sci-vibes-retalt1"
VOLUME_NAME = "sci-vibes-retalt1-v1"
MOUNT = Path("/mnt/dataset")
WORKBOOK_URL = (
    "https://zenodo.org/api/records/7027367/files/"
    "RETALT1_DLR_CFSE_AEDB2.0_PlanarFins%2CPetals_2022-07-25.xlsx/content"
)
WORKBOOK_MD5 = "cd79b335a671e50e0ac2b177d6b604c2"

META_SHEETS = {"Issue", "Reference values", "Configurations", "Uncertainties"}
TARGETS = ["CA", "CFy", "CN", "CMx", "CMy", "CMz", "CMy_cog", "CMz_cog"]
TOLERANCES = ["UCA", "UCFy", "UCN", "UCMx", "UCMy", "UCMz", "UCMy_cog", "UCMz_cog"]

# This list is intentionally available only to the manager-side Modal job. The
# branch repositories receive a blinded row count, not these names or labels.
PRIVATE_CONFIGURATION_SHEETS = {
    "PF10,0,10,0_Pitch",
    "PF0,10,0,10_Yaw",
    "PF10,10,-10-10_Roll",
    "PF0,10,0,0_Shadow",
    "B10,10,10,10 Drag",
    "B0,10,0,0 Pitch",
    "B10,10,0,0 XPitch",
}

image = modal.Image.debian_slim(python_version="3.12").pip_install(
    "openpyxl==3.1.5",
    "pandas==2.2.3",
)

app = modal.App(APP_NAME)
volume = modal.Volume.from_name(VOLUME_NAME, create_if_missing=True)


def _layout(sheet_name: str) -> str:
    lowered = sheet_name.lower().replace("-", "_")
    for needle, value in (
        ("x_shadow", "x_shadow"),
        ("x_roll", "x_roll"),
        ("x_pitch", "x_pitch"),
        ("xpitch", "x_pitch"),
        ("shadow", "shadow"),
        ("roll", "roll"),
        ("yaw", "yaw"),
        ("pitch", "pitch"),
        ("drag", "drag"),
    ):
        if needle in lowered:
            return value
    return "baseline"


def _deflections(sheet_name: str) -> list[int]:
    prefix = re.split(r"[_ ]", sheet_name, maxsplit=1)[0]
    values = [int(value) for value in re.findall(r"-?\d+", prefix)]
    if len(values) != 4:
        raise ValueError(f"Could not parse four control-surface deflections from {sheet_name!r}.")
    return values


def _download_workbook(target: Path) -> None:
    target.parent.mkdir(parents=True, exist_ok=True)
    with urlopen(WORKBOOK_URL, timeout=120) as response:
        target.write_bytes(response.read())
    digest = hashlib.md5(target.read_bytes(), usedforsecurity=False).hexdigest()
    if digest != WORKBOOK_MD5:
        raise ValueError(f"RETALT1 checksum mismatch: expected {WORKBOOK_MD5}, got {digest}.")


def _run_deflections(run_id: str) -> list[int] | None:
    match = re.search(
        r"_(?:X_Pitch|X_Roll|Pitch|Yaw|Roll|Shadow|Drag)_"
        r"(-?\d+),(-?\d+),(-?\d+),(-?\d+)_",
        run_id,
    )
    return [int(value) for value in match.groups()] if match else None


def _normalize_sheet(pandas, worksheet):
    sheet_name = worksheet.title
    rows = worksheet.iter_rows(values_only=True)
    headers = list(next(rows))
    while headers and headers[-1] is None:
        headers.pop()

    required = {"Run", "MACH", "ALPHA DSC", *TARGETS, *TOLERANCES}
    missing = sorted(required - set(headers))
    if missing:
        raise ValueError(f"{sheet_name}: missing common-schema columns {missing}")

    sheet_deltas = _deflections(sheet_name)
    normalized = []
    for excel_row, values in enumerate(rows, start=2):
        record = dict(zip(headers, values[: len(headers)]))
        run_id = record.get("Run")
        if run_id in (None, ""):
            continue
        # The 5-degree planar-fin pitch sheet includes a legitimate extra
        # negative-deflection run. Elsewhere the sheet declaration is the
        # canonical source because two run labels are known to be wrong.
        deltas = sheet_deltas
        if sheet_name == "PF5,0,5,0_Pitch" and str(run_id).startswith("PF_Pitch_-05"):
            deltas = _run_deflections(str(run_id)) or sheet_deltas
        normalized.append(
            {
                "curve_id": f"{sheet_name}::{run_id}",
                "raw_run_id": run_id,
                "source_sheet": sheet_name,
                "excel_row": excel_row,
                "source_kind": "aedb",
                "surface_family": "planar_fin" if sheet_name.startswith("PF") else "petal",
                "layout": _layout(sheet_name),
                "engine_state": "UFN",
                "delta_1_deg": deltas[0],
                "delta_2_deg": deltas[1],
                "delta_3_deg": deltas[2],
                "delta_4_deg": deltas[3],
                "mach": float(record["MACH"]),
                "alpha_deg": float(record["ALPHA DSC"]),
                **{name: record[name] for name in [*TARGETS, *TOLERANCES]},
            }
        )
    frame = pandas.DataFrame(normalized)
    ordered = [
        "curve_id",
        "raw_run_id",
        "source_sheet",
        "excel_row",
        "source_kind",
        "surface_family",
        "layout",
        "engine_state",
        "delta_1_deg",
        "delta_2_deg",
        "delta_3_deg",
        "delta_4_deg",
        "mach",
        "alpha_deg",
        *TARGETS,
        *TOLERANCES,
    ]
    return frame[ordered]


@app.function(
    image=image,
    volumes={str(MOUNT): volume},
    timeout=15 * 60,
)
def prepare_dataset() -> dict:
    import pandas
    from openpyxl import load_workbook

    raw_path = MOUNT / "private" / "raw" / "retalt1-aedb-v3.xlsx"
    _download_workbook(raw_path)

    workbook = load_workbook(raw_path, read_only=True, data_only=True)
    coefficient_sheets = [
        name
        for name in workbook.sheetnames
        if name not in META_SHEETS and "CFD NS" not in name
    ]
    frames = [_normalize_sheet(pandas, workbook[name]) for name in coefficient_sheets]
    dataset = pandas.concat(frames, ignore_index=True)

    is_hidden = dataset["source_sheet"].isin(PRIVATE_CONFIGURATION_SHEETS)
    hidden = dataset.loc[is_hidden].copy()
    development = dataset.loc[~is_hidden].copy()
    is_public_validation = development["mach"].eq(2.5)
    validation = development.loc[is_public_validation].copy()
    train = development.loc[~is_public_validation].copy()

    expected = {"all": 6908, "train": 4408, "validation": 779, "hidden": 1721}
    actual = {
        "all": int(len(dataset)),
        "train": int(len(train)),
        "validation": int(len(validation)),
        "hidden": int(len(hidden)),
    }
    if actual != expected:
        raise ValueError(f"RETALT1 split drifted. Expected {expected}, got {actual}.")

    public_dir = MOUNT / "public"
    private_dir = MOUNT / "private"
    public_dir.mkdir(parents=True, exist_ok=True)
    private_dir.mkdir(parents=True, exist_ok=True)

    train.to_csv(public_dir / "train.csv", index=False)
    validation.to_csv(public_dir / "validation.csv", index=False)
    hidden.to_csv(private_dir / "hidden.csv", index=False)
    file_hashes = {
        "train_sha256": hashlib.sha256(
            (public_dir / "train.csv").read_bytes()
        ).hexdigest(),
        "validation_sha256": hashlib.sha256(
            (public_dir / "validation.csv").read_bytes()
        ).hexdigest(),
        "sealed_hidden_sha256": hashlib.sha256(
            (private_dir / "hidden.csv").read_bytes()
        ).hexdigest(),
    }

    public_manifest = {
        "dataset_id": "retalt1-aedb-v3",
        "name": "RETALT1 Aerodynamic Data Base 2.0",
        "version": "v3",
        "source": "https://zenodo.org/records/7027367",
        "doi": "10.5281/zenodo.7027367",
        "license": "CC BY 4.0",
        "workbook_md5": WORKBOOK_MD5,
        "file_hashes": file_hashes,
        "rows": actual,
        "curves": {
            "all": int(dataset["curve_id"].nunique()),
            "train": int(train["curve_id"].nunique()),
            "validation": int(validation["curve_id"].nunique()),
            "hidden": int(hidden["curve_id"].nunique()),
        },
        "features": [
            "surface_family",
            "layout",
            "engine_state",
            "delta_1_deg",
            "delta_2_deg",
            "delta_3_deg",
            "delta_4_deg",
            "mach",
            "alpha_deg",
        ],
        "targets": TARGETS,
        "tolerances": TOLERANCES,
        "split": {
            "train": "Development configurations excluding Mach 2.5.",
            "validation": "Mach 2.5 curves from development configurations.",
            "hidden": "Seven complete control-surface configurations; manager-only labels.",
        },
        "scientific_scope": (
            "Processed integral aerodynamic force and moment coefficients. "
            "This is not a turbulence-field or CFD-mesh dataset."
        ),
    }
    private_manifest = {
        **public_manifest,
        "private_configuration_sheets": sorted(PRIVATE_CONFIGURATION_SHEETS),
        "private_files": ["private/raw/retalt1-aedb-v3.xlsx", "private/hidden.csv"],
    }

    (public_dir / "manifest.json").write_text(
        json.dumps(public_manifest, indent=2) + "\n", encoding="utf-8"
    )
    (private_dir / "split-manifest.json").write_text(
        json.dumps(private_manifest, indent=2) + "\n", encoding="utf-8"
    )
    volume.commit()
    return {**public_manifest, "modal_volume": VOLUME_NAME, "status": "prepared"}


@app.function(
    image=image,
    volumes={str(MOUNT): volume.with_mount_options(read_only=True)},
    block_network=True,
    timeout=5 * 60,
)
def hidden_dataset_summary() -> dict:
    """Manager-only integrity check that never returns labels or configuration names."""

    import pandas

    hidden = pandas.read_csv(MOUNT / "private" / "hidden.csv")
    return {
        "rows": int(len(hidden)),
        "curves": int(hidden["curve_id"].nunique()),
        "targets": TARGETS,
        "sha256": hashlib.sha256(
            (MOUNT / "private" / "hidden.csv").read_bytes()
        ).hexdigest(),
    }


@app.local_entrypoint()
def main() -> None:
    prepared = prepare_dataset.remote()
    integrity = hidden_dataset_summary.remote()
    print(json.dumps({"prepared": prepared, "hidden_integrity": integrity}, indent=2))
